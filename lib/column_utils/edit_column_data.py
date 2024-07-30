from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.exceptions import InvalidFileException

def edit_column_data_from_excel(xlsx_path: str, data: list, row: int = 1, column: str = 'A', hyperlink:bool = False) -> bool:
    """
    Fills a column in an Excel file with provided data.
    
    :param excel_path: Path to the Excel file.
    :param data: List of data to fill. Can be [(data, color), (data, color)...], [data, data...] or [(data, color), data...]
    :param row: Starting row for filling (1-indexed).
    :param column: Column to fill.
    :param hyperlink: Whether to add hyperlinks to the cells.
    :raises ValueError: If the specified row or column does not exist.
    :raises InvalidFileException: If the provided path does not lead to a valid .xlsx file.
    """
    try:
        load_wb = load_workbook(xlsx_path)
        load_ws = load_wb.active
        
        column = column.upper()
        row = row - 1
        for i, item in enumerate(data, start=row):
            cell = load_ws[f'{column}{i+1}']
            if isinstance(item, tuple):
                value, color = item
                cell.value = value
                if color:
                    cell.fill = PatternFill(
                        start_color=color,
                        end_color=color,
                        fill_type='solid'
                    )
                if hyperlink:
                    cell.hyperlink = value
            else:
                cell.value = item
                if hyperlink:
                    cell.hyperlink = item
        
        load_wb.save(xlsx_path)
    
    except ValueError:
        raise ValueError("The specified row or column does not exist.")

    except InvalidFileException:
        raise InvalidFileException("The path does not lead to a valid .xlsx file.")