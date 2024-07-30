import pytest

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

import os
import sys

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from lib.column_utils.get_column_data  import get_column_data_from_excel
from lib.column_utils.edit_column_data import edit_column_data_from_excel

@pytest.fixture
def sample_excel_file(tmp_path):
    workbook = Workbook()

    file_path = tmp_path / 'sample2.xlsx'
    workbook.save(file_path)
    return file_path

def test_edit_column_data_from_excel_valid(sample_excel_file):
    edit_column_data_from_excel(
        xlsx_path=str(sample_excel_file),
        data=['Header', 'Row 1', None, ('Row 3', 'ffcccb')],
        row=1,
        column='A'
    )

    result = get_column_data_from_excel(
        xlsx_path=str(sample_excel_file)
    )

    expected = ['Header', 'Row 1', 'Row 3']
    assert result == expected

def test_edit_column_data_from_excel_with_index(sample_excel_file):
    edit_column_data_from_excel(
        xlsx_path=str(sample_excel_file),
        data=['Header', 'Row 1', None, ('Row 3', 'ffcccb')],
        row=1,
        column='A'
    )

    result = get_column_data_from_excel(
        xlsx_path=str(sample_excel_file),
        index=True
    )

    expected = [(1, 'Header'), (2, 'Row 1'), (4, 'Row 3')]
    assert result == expected

def test_edit_column_data_from_excel_cell_color(sample_excel_file):
    edit_column_data_from_excel(
        xlsx_path=str(sample_excel_file),
        data=[('Colored Cell', 'ffcccb')],
        row=1,
        column='A'
    )

    workbook = load_workbook(sample_excel_file)
    sheet = workbook.active

    cell = sheet['A1']
    assert cell.value == 'Colored Cell'
    assert cell.fill.start_color.index == '00ffcccb' # prefix 00 is the value of transparency
    
def test_edit_column_data_from_excel_cell_hyperlink(sample_excel_file):
    edit_column_data_from_excel(
        xlsx_path=str(sample_excel_file),
        data=[('http://example.com', None)],
        row=1,
        column='A',
        hyperlink=True
    )

    workbook = load_workbook(sample_excel_file)
    sheet = workbook.active

    cell = sheet['A1']
    assert cell.value == 'http://example.com'
    assert cell.hyperlink.target == 'http://example.com'

def test_edit_column_data_from_excel_invalid_file(tmp_path):
    invalid_file = tmp_path / 'invalid_file.txt'
    invalid_file.write_text('This is not an Excel file.')

    with pytest.raises(InvalidFileException):
        get_column_data_from_excel(str(invalid_file))
    
def test_edit_column_data_from_excel_invalid_column(sample_excel_file):
    with pytest.raises(ValueError):
        edit_column_data_from_excel(
            str(sample_excel_file),
            data=[None],
            row=1,
            column='ABZZ'
        )

def test_edit_column_data_from_excel_invalid_row(sample_excel_file):
    with pytest.raises(ValueError):
        edit_column_data_from_excel(
            str(sample_excel_file),
            data=[None],
            row=100000000,
            column='A'
        )